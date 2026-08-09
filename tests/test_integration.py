"""
Integration tests: a few important end-to-end workflows exercised through
the real HTTP stack (FastAPI TestClient), verifying both the API responses
and the resulting database state directly via db_session.
"""
import io
from datetime import datetime, timedelta

import openpyxl

from app.core.constants import (
    AuditAction,
    ReservationStatus,
    RoleName,
    SetupStatus,
    SwapStatus,
)
from app.models.audit_log import AuditLog
from app.models.reservation import Reservation
from app.models.setup import Setup

API = "/api/v1"


def _iso(dt):
    return dt.isoformat()


def _window():
    start = datetime.utcnow() + timedelta(hours=1)
    return start, start + timedelta(hours=2)


# ---------------------------------------------------------------------
# 1. Register -> Login -> Reserve -> verify DB/log
# ---------------------------------------------------------------------

def test_workflow_register_login_reserve_verify_db_and_log(client, db_session, make_user, auth_headers, setup):
    register_resp = client.post(
        "{0}/auth/register".format(API),
        json={
            "username": "flowuser1",
            "email": "flowuser1@example.com",
            "password": "Password123",
            "full_name": "Flow User One",
        },
    )
    assert register_resp.status_code == 201
    new_user_id = register_resp.json()["id"]

    # Registration lands in PENDING -- login must fail until approved.
    login_before_approval = client.post(
        "{0}/auth/login".format(API), json={"username": "flowuser1", "password": "Password123"}
    )
    assert login_before_approval.status_code == 401

    approver = make_user(role_name=RoleName.DEVELOPER_LEAD)
    approve_resp = client.post(
        "{0}/users/{1}/approval".format(API, new_user_id),
        json={"approve": True, "role_name": RoleName.DEVELOPER},
        headers=auth_headers(approver),
    )
    assert approve_resp.status_code == 200

    login_resp = client.post(
        "{0}/auth/login".format(API), json={"username": "flowuser1", "password": "Password123"}
    )
    assert login_resp.status_code == 200
    access_token = login_resp.json()["access_token"]
    headers = {"Authorization": "Bearer {0}".format(access_token)}

    start, end = _window()
    reserve_resp = client.post(
        "{0}/reservations".format(API),
        json={"setup_id": setup.id, "reserved_from": _iso(start), "reserved_until": _iso(end)},
        headers=headers,
    )
    assert reserve_resp.status_code == 201
    reservation_id = reserve_resp.json()["id"]

    # Verify DB state directly.
    db_reservation = db_session.query(Reservation).filter(Reservation.id == reservation_id).first()
    assert db_reservation is not None
    assert db_reservation.user_id == new_user_id
    assert db_reservation.status == ReservationStatus.ACTIVE

    db_setup = db_session.query(Setup).filter(Setup.id == setup.id).first()
    assert db_setup.status == SetupStatus.RESERVED

    # Verify the audit trail recorded both the registration and the reservation.
    create_user_logs = (
        db_session.query(AuditLog)
        .filter(AuditLog.entity_type == "User", AuditLog.entity_id == new_user_id, AuditLog.action == AuditAction.CREATE)
        .all()
    )
    assert len(create_user_logs) >= 1

    reservation_logs = (
        db_session.query(AuditLog)
        .filter(AuditLog.entity_type == "Reservation", AuditLog.entity_id == reservation_id, AuditLog.action == AuditAction.CREATE)
        .all()
    )
    assert len(reservation_logs) == 1


# ---------------------------------------------------------------------
# 2. Reserve -> Swap -> verify DB/history
# ---------------------------------------------------------------------

def test_workflow_reserve_swap_verify_history(client, db_session, auth_headers, developer_user, make_user, make_setup, product):
    dev_headers = auth_headers(developer_user)
    setup_a = make_setup(product_id=product.id)
    setup_b = make_setup(product_id=product.id)
    start, end = _window()

    reserve_resp = client.post(
        "{0}/reservations".format(API),
        json={"setup_id": setup_a.id, "reserved_from": _iso(start), "reserved_until": _iso(end)},
        headers=dev_headers,
    )
    assert reserve_resp.status_code == 201
    original_reservation_id = reserve_resp.json()["id"]

    swap_resp = client.post(
        "{0}/swaps".format(API),
        json={"reservation_id": original_reservation_id, "requested_setup_id": setup_b.id},
        headers=dev_headers,
    )
    assert swap_resp.status_code == 201
    swap_id = swap_resp.json()["id"]

    approver = make_user(role_name=RoleName.LEAD)
    approve_resp = client.patch(
        "{0}/swaps/{1}/approve".format(API, swap_id), json={}, headers=auth_headers(approver)
    )
    assert approve_resp.status_code == 200
    assert approve_resp.json()["status"] == SwapStatus.COMPLETED

    # Original reservation is now SWAPPED (terminal), setup A freed, setup B reserved.
    db_session.expire_all()
    original = db_session.query(Reservation).filter(Reservation.id == original_reservation_id).first()
    assert original.status == ReservationStatus.SWAPPED
    assert original.remarks is not None
    assert developer_user.email in original.remarks
    assert setup_a.hostname in original.remarks
    assert setup_b.hostname in original.remarks

    new_reservation = (
        db_session.query(Reservation)
        .filter(Reservation.setup_id == setup_b.id, Reservation.status == ReservationStatus.ACTIVE)
        .first()
    )
    assert new_reservation is not None
    assert new_reservation.user_id == developer_user.id
    assert new_reservation.remarks == original.remarks  # history carried forward

    db_setup_a = db_session.query(Setup).filter(Setup.id == setup_a.id).first()
    db_setup_b = db_session.query(Setup).filter(Setup.id == setup_b.id).first()
    assert db_setup_a.status == SetupStatus.AVAILABLE
    assert db_setup_b.status == SetupStatus.RESERVED


# ---------------------------------------------------------------------
# 3. Active swap -> Unreserve -> verify rejection
# ---------------------------------------------------------------------

def test_workflow_active_swap_blocks_unreserve(client, auth_headers, developer_user, make_setup, product):
    dev_headers = auth_headers(developer_user)
    setup_a = make_setup(product_id=product.id)
    setup_b = make_setup(product_id=product.id)
    start, end = _window()

    reserve_resp = client.post(
        "{0}/reservations".format(API),
        json={"setup_id": setup_a.id, "reserved_from": _iso(start), "reserved_until": _iso(end)},
        headers=dev_headers,
    )
    reservation_id = reserve_resp.json()["id"]

    swap_resp = client.post(
        "{0}/swaps".format(API),
        json={"reservation_id": reservation_id, "requested_setup_id": setup_b.id},
        headers=dev_headers,
    )
    assert swap_resp.status_code == 201

    cancel_resp = client.patch("{0}/reservations/{1}/cancel".format(API, reservation_id), headers=dev_headers)
    assert cancel_resp.status_code == 409
    assert cancel_resp.json()["error"]["code"] == "CONFLICT"


# ---------------------------------------------------------------------
# 4. Restore swap -> Unreserve -> verify success
# ---------------------------------------------------------------------

def test_workflow_restored_swap_allows_unreserve(client, db_session, auth_headers, developer_user, make_setup, product):
    dev_headers = auth_headers(developer_user)
    setup_a = make_setup(product_id=product.id)
    setup_b = make_setup(product_id=product.id)
    start, end = _window()

    reserve_resp = client.post(
        "{0}/reservations".format(API),
        json={"setup_id": setup_a.id, "reserved_from": _iso(start), "reserved_until": _iso(end)},
        headers=dev_headers,
    )
    reservation_id = reserve_resp.json()["id"]

    swap_resp = client.post(
        "{0}/swaps".format(API),
        json={"reservation_id": reservation_id, "requested_setup_id": setup_b.id},
        headers=dev_headers,
    )
    swap_id = swap_resp.json()["id"]

    cancel_swap_resp = client.patch("{0}/swaps/{1}/cancel".format(API, swap_id), headers=dev_headers)
    assert cancel_swap_resp.status_code == 200
    assert cancel_swap_resp.json()["status"] == SwapStatus.CANCELLED

    unreserve_resp = client.patch("{0}/reservations/{1}/cancel".format(API, reservation_id), headers=dev_headers)
    assert unreserve_resp.status_code == 200
    assert unreserve_resp.json()["status"] == ReservationStatus.CANCELLED

    db_session.expire_all()
    db_setup_a = db_session.query(Setup).filter(Setup.id == setup_a.id).first()
    assert db_setup_a.status == SetupStatus.AVAILABLE


# ---------------------------------------------------------------------
# 5. Two users -> same setup -> verify only one reservation
# ---------------------------------------------------------------------

def test_workflow_two_users_same_setup_only_one_reservation_wins(
    client, db_session, auth_headers, developer_user, second_developer_user, setup
):
    start, end = _window()
    payload = {"setup_id": setup.id, "reserved_from": _iso(start), "reserved_until": _iso(end)}

    first_resp = client.post("{0}/reservations".format(API), json=payload, headers=auth_headers(developer_user))
    second_resp = client.post(
        "{0}/reservations".format(API), json=payload, headers=auth_headers(second_developer_user)
    )

    assert first_resp.status_code == 201
    assert second_resp.status_code == 409

    active_reservations = (
        db_session.query(Reservation)
        .filter(Reservation.setup_id == setup.id, Reservation.status == ReservationStatus.ACTIVE)
        .all()
    )
    assert len(active_reservations) == 1
    assert active_reservations[0].user_id == developer_user.id


# ---------------------------------------------------------------------
# 6. Product -> Import Excel -> Export -> verify data
# ---------------------------------------------------------------------

def test_workflow_product_import_then_export_roundtrip(client, auth_headers, owner_user, product):
    from app.utils.excel_reader import SETUP_IMPORT_COLUMNS

    headers = auth_headers(owner_user)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(SETUP_IMPORT_COLUMNS)
    row = {
        "product_name": product.name,
        "ip_address": "10.9.9.9",
        "hostname": "roundtrip-host.example.com",
        "location": "Rack RT",
        "remarks": "Imported by integration test",
    }
    sheet.append([row.get(col, "") for col in SETUP_IMPORT_COLUMNS])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    import_resp = client.post(
        "{0}/imports/setups".format(API),
        files={"file": ("setups.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert import_resp.status_code == 200
    import_body = import_resp.json()
    assert import_body["committed"] is True
    assert import_body["created_count"] == 1

    export_resp = client.post(
        "{0}/exports/setups?product_id={1}".format(API, product.id), headers=headers
    )
    assert export_resp.status_code == 200

    exported_workbook = openpyxl.load_workbook(io.BytesIO(export_resp.content))
    sheet_out = exported_workbook["Setups"]
    header_row = [cell.value for cell in sheet_out[1]]
    data_rows = list(sheet_out.iter_rows(min_row=2, values_only=True))

    assert "hostname" in header_row
    assert "ip_address" in header_row
    hostname_index = header_row.index("hostname")
    ip_index = header_row.index("ip_address")

    matching = [r for r in data_rows if r[hostname_index] == "roundtrip-host.example.com"]
    assert len(matching) == 1
    assert matching[0][ip_index] == "10.9.9.9"
