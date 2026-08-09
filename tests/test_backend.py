"""
Backend/API tests: registration, login, auth, RBAC, approval, Product CRUD,
Excel import/export, reservation/swap/unreserve basics, announcements,
audit logging.

These go through the real HTTP stack (FastAPI TestClient), one independent
request per call, exactly as a real client would -- this is deliberate so
that any gap between "the data was flushed" and "the data was durably
committed and visible to the next request" is caught rather than hidden.
"""
import io
from datetime import datetime, timedelta

import openpyxl

from app.core.constants import RoleName, SetupStatus, UserStatus

API = "/api/v1"


# ---------------------------------------------------------------------
# Registration
# ---------------------------------------------------------------------

def test_register_creates_pending_user(client):
    response = client.post(
        "{0}/auth/register".format(API),
        json={
            "username": "newbie1",
            "email": "newbie1@example.com",
            "password": "Password123",
            "full_name": "New Bie",
        },
    )
    assert response.status_code == 201
    body = response.json()
    assert body["status"] == UserStatus.PENDING
    assert body["username"] == "newbie1"


def test_register_duplicate_username_rejected(client):
    payload = {
        "username": "dupeuser",
        "email": "dupe1@example.com",
        "password": "Password123",
        "full_name": "Dupe One",
    }
    first = client.post("{0}/auth/register".format(API), json=payload)
    assert first.status_code == 201

    payload["email"] = "dupe2@example.com"
    second = client.post("{0}/auth/register".format(API), json=payload)
    assert second.status_code == 409
    assert second.json()["error"]["code"] == "CONFLICT"


def test_register_weak_password_rejected(client):
    response = client.post(
        "{0}/auth/register".format(API),
        json={
            "username": "weakpass",
            "email": "weakpass@example.com",
            "password": "alllowercase",
            "full_name": "Weak Pass",
        },
    )
    assert response.status_code == 422


# ---------------------------------------------------------------------
# Login / Authentication
# ---------------------------------------------------------------------

def test_login_success_returns_token_pair(client, make_user):
    make_user(username="loginok", password="Password123", status=UserStatus.APPROVED)
    response = client.post("{0}/auth/login".format(API), json={"username": "loginok", "password": "Password123"})
    assert response.status_code == 200
    body = response.json()
    assert "access_token" in body and "refresh_token" in body
    assert body["token_type"] == "bearer"


def test_login_wrong_password_rejected(client, make_user):
    make_user(username="loginbad", password="Password123")
    response = client.post("{0}/auth/login".format(API), json={"username": "loginbad", "password": "WrongPass1"})
    assert response.status_code == 401


def test_login_pending_user_rejected(client, make_user):
    make_user(username="pendinglogin", password="Password123", status=UserStatus.PENDING)
    response = client.post("{0}/auth/login".format(API), json={"username": "pendinglogin", "password": "Password123"})
    assert response.status_code == 401
    assert response.json()["error"]["code"] == "ACCOUNT_NOT_APPROVED"


def test_login_disabled_user_rejected(client, make_user):
    make_user(username="disabledlogin", password="Password123", status=UserStatus.DISABLED, is_active=False)
    response = client.post("{0}/auth/login".format(API), json={"username": "disabledlogin", "password": "Password123"})
    assert response.status_code == 401
    assert response.json()["error"]["code"] == "ACCOUNT_DISABLED"


def test_authenticated_endpoint_without_token_rejected(client):
    response = client.get("{0}/users/me".format(API))
    assert response.status_code == 401


def test_users_me_returns_current_user(client, auth_headers, developer_user):
    response = client.get("{0}/users/me".format(API), headers=auth_headers(developer_user))
    assert response.status_code == 200
    assert response.json()["username"] == developer_user.username


def test_data_written_in_one_request_persists_and_is_visible_in_a_separate_request(client):
    """
    Regression test for a session-persistence bug: ``get_db()`` must commit
    the request-scoped session on success. Each TestClient call is an
    independent HTTP request (a fresh SQLAlchemy session via ``get_db``), so
    if a write is only flushed (not committed) within its own request, the
    session close() at the end of that request silently rolls it back and
    it would never be visible to this second, independent request.
    """
    create_response = client.post(
        "{0}/auth/register".format(API),
        json={
            "username": "persistencecheck",
            "email": "persistencecheck@example.com",
            "password": "Password123",
            "full_name": "Persistence Check",
        },
    )
    assert create_response.status_code == 201

    duplicate_response = client.post(
        "{0}/auth/register".format(API),
        json={
            "username": "persistencecheck",
            "email": "different@example.com",
            "password": "Password123",
            "full_name": "Persistence Check",
        },
    )
    assert duplicate_response.status_code == 409, (
        "Expected the username created in the first request to be visible to this second, "
        "independent request. A 201 here means the first request's write was never committed."
    )


# ---------------------------------------------------------------------
# RBAC
# ---------------------------------------------------------------------

def test_plain_user_cannot_create_product(client, auth_headers, plain_user):
    response = client.post(
        "{0}/products".format(API), json={"name": "Blocked Product"}, headers=auth_headers(plain_user)
    )
    assert response.status_code == 403
    assert response.json()["error"]["code"] == "AUTHORIZATION_ERROR"


def test_owner_can_create_product(client, auth_headers, owner_user):
    response = client.post(
        "{0}/products".format(API), json={"name": "Owner Product"}, headers=auth_headers(owner_user)
    )
    assert response.status_code == 201


def test_developer_cannot_approve_users(client, auth_headers, developer_user, make_user):
    pending = make_user(status=UserStatus.PENDING)
    response = client.post(
        "{0}/users/{1}/approval".format(API, pending.id), json={"approve": True}, headers=auth_headers(developer_user)
    )
    assert response.status_code == 403


def test_developer_cannot_view_audit_log(client, auth_headers, developer_user):
    response = client.get("{0}/audit-logs".format(API), headers=auth_headers(developer_user))
    assert response.status_code == 403


def test_developer_lead_can_view_audit_log(client, auth_headers, make_user):
    dev_lead = make_user(role_name=RoleName.DEVELOPER_LEAD)
    response = client.get("{0}/audit-logs".format(API), headers=auth_headers(dev_lead))
    assert response.status_code == 200


# ---------------------------------------------------------------------
# Role approval workflow
# ---------------------------------------------------------------------

def test_lead_can_approve_user_in_same_group(client, auth_headers, make_user, group):
    lead = make_user(role_name=RoleName.LEAD, group_id=group.id)
    pending = make_user(status=UserStatus.PENDING, group_id=group.id)

    response = client.post(
        "{0}/users/{1}/approval".format(API, pending.id),
        json={"approve": True, "role_name": RoleName.DEVELOPER},
        headers=auth_headers(lead),
    )
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == UserStatus.APPROVED
    assert body["role"]["name"] == RoleName.DEVELOPER


def test_lead_cannot_approve_user_outside_group(client, auth_headers, make_user, group):
    other_group_lead = make_user(role_name=RoleName.LEAD, group_id=group.id)
    pending_in_other_group = make_user(status=UserStatus.PENDING, group_id=None)

    response = client.post(
        "{0}/users/{1}/approval".format(API, pending_in_other_group.id),
        json={"approve": True},
        headers=auth_headers(other_group_lead),
    )
    assert response.status_code == 403


def test_reject_sets_status_rejected(client, auth_headers, make_user):
    dev_lead = make_user(role_name=RoleName.DEVELOPER_LEAD)
    pending = make_user(status=UserStatus.PENDING)

    response = client.post(
        "{0}/users/{1}/approval".format(API, pending.id),
        json={"approve": False, "rejection_reason": "Not needed"},
        headers=auth_headers(dev_lead),
    )
    assert response.status_code == 200
    assert response.json()["status"] == UserStatus.REJECTED


def test_approving_already_approved_user_conflicts(client, auth_headers, make_user):
    dev_lead = make_user(role_name=RoleName.DEVELOPER_LEAD)
    already_approved = make_user(status=UserStatus.APPROVED)

    response = client.post(
        "{0}/users/{1}/approval".format(API, already_approved.id),
        json={"approve": True},
        headers=auth_headers(dev_lead),
    )
    assert response.status_code == 409


# ---------------------------------------------------------------------
# Product CRUD
# ---------------------------------------------------------------------

def test_product_crud_lifecycle(client, auth_headers, owner_user):
    headers = auth_headers(owner_user)

    create = client.post("{0}/products".format(API), json={"name": "CRUD Product"}, headers=headers)
    assert create.status_code == 201
    product_id = create.json()["id"]

    get_resp = client.get("{0}/products/{1}".format(API, product_id), headers=headers)
    assert get_resp.status_code == 200
    assert get_resp.json()["name"] == "CRUD Product"

    update = client.patch(
        "{0}/products/{1}".format(API, product_id), json={"description": "Updated"}, headers=headers
    )
    assert update.status_code == 200
    assert update.json()["description"] == "Updated"

    delete = client.delete("{0}/products/{1}".format(API, product_id), headers=headers)
    assert delete.status_code == 200

    get_after_delete = client.get("{0}/products/{1}".format(API, product_id), headers=headers)
    assert get_after_delete.status_code == 404


def test_product_duplicate_name_rejected(client, auth_headers, owner_user):
    headers = auth_headers(owner_user)
    client.post("{0}/products".format(API), json={"name": "Unique Product"}, headers=headers)
    response = client.post("{0}/products".format(API), json={"name": "Unique Product"}, headers=headers)
    assert response.status_code == 409


def test_product_delete_blocked_while_setup_assigned(client, auth_headers, owner_user, product, setup):
    response = client.delete("{0}/products/{1}".format(API, product.id), headers=auth_headers(owner_user))
    assert response.status_code == 409


# ---------------------------------------------------------------------
# Excel import/export
# ---------------------------------------------------------------------

def _build_setup_import_workbook(rows):
    from app.utils.excel_reader import SETUP_IMPORT_COLUMNS

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(SETUP_IMPORT_COLUMNS)
    for row in rows:
        sheet.append([row.get(col, "") for col in SETUP_IMPORT_COLUMNS])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def test_import_setups_success(client, auth_headers, owner_user, product):
    buffer = _build_setup_import_workbook(
        [{"product_name": product.name, "ip_address": "10.5.5.5", "hostname": "import-host.example.com", "location": "Rack Z"}]
    )
    response = client.post(
        "{0}/imports/setups".format(API),
        files={"file": ("setups.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers(owner_user),
    )
    assert response.status_code == 200
    body = response.json()
    assert body["committed"] is True
    assert body["created_count"] == 1
    assert body["error_count"] == 0


def test_import_setups_unknown_product_rejected(client, auth_headers, owner_user):
    buffer = _build_setup_import_workbook(
        [{"product_name": "Does Not Exist", "ip_address": "10.5.5.6", "hostname": "bad-host.example.com", "location": "Rack Z"}]
    )
    response = client.post(
        "{0}/imports/setups".format(API),
        files={"file": ("setups.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers(owner_user),
    )
    assert response.status_code == 200
    body = response.json()
    assert body["committed"] is False
    assert body["error_count"] >= 1


def test_export_setups_returns_xlsx(client, auth_headers, owner_user, setup):
    response = client.post("{0}/exports/setups".format(API), headers=auth_headers(owner_user))
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert "Setups" in workbook.sheetnames


def test_export_setups_empty_generates_template(client, auth_headers, owner_user, product):
    response = client.post(
        "{0}/exports/setups?product_id={1}".format(API, product.id), headers=auth_headers(owner_user)
    )
    assert response.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert "Template" in workbook.sheetnames


# ---------------------------------------------------------------------
# Reservation
# ---------------------------------------------------------------------

def _iso(dt):
    return dt.isoformat()


def test_create_reservation_success(client, auth_headers, developer_user, setup):
    start = datetime.utcnow() + timedelta(hours=1)
    end = start + timedelta(hours=2)
    response = client.post(
        "{0}/reservations".format(API),
        json={"setup_id": setup.id, "reserved_from": _iso(start), "reserved_until": _iso(end)},
        headers=auth_headers(developer_user),
    )
    assert response.status_code == 201
    body = response.json()
    assert body["setup_id"] == setup.id
    assert body["user_id"] == developer_user.id
    assert body["status"] == "ACTIVE"


def test_multiple_reservations_different_setups(client, auth_headers, developer_user, make_setup):
    setup_a = make_setup()
    setup_b = make_setup()
    start = datetime.utcnow() + timedelta(hours=1)
    end = start + timedelta(hours=2)
    headers = auth_headers(developer_user)

    resp_a = client.post(
        "{0}/reservations".format(API),
        json={"setup_id": setup_a.id, "reserved_from": _iso(start), "reserved_until": _iso(end)},
        headers=headers,
    )
    resp_b = client.post(
        "{0}/reservations".format(API),
        json={"setup_id": setup_b.id, "reserved_from": _iso(start), "reserved_until": _iso(end)},
        headers=headers,
    )
    assert resp_a.status_code == 201
    assert resp_b.status_code == 201


def test_reservation_conflict_overlapping_window(client, auth_headers, developer_user, second_developer_user, setup):
    start = datetime.utcnow() + timedelta(hours=1)
    end = start + timedelta(hours=2)

    first = client.post(
        "{0}/reservations".format(API),
        json={"setup_id": setup.id, "reserved_from": _iso(start), "reserved_until": _iso(end)},
        headers=auth_headers(developer_user),
    )
    assert first.status_code == 201

    second = client.post(
        "{0}/reservations".format(API),
        json={"setup_id": setup.id, "reserved_from": _iso(start), "reserved_until": _iso(end)},
        headers=auth_headers(second_developer_user),
    )
    assert second.status_code == 409
    assert second.json()["error"]["code"] == "RESERVATION_CONFLICT"


def test_reservation_invalid_window_rejected(client, auth_headers, developer_user, setup):
    start = datetime.utcnow() + timedelta(hours=2)
    end = start - timedelta(hours=1)  # until before from
    response = client.post(
        "{0}/reservations".format(API),
        json={"setup_id": setup.id, "reserved_from": _iso(start), "reserved_until": _iso(end)},
        headers=auth_headers(developer_user),
    )
    assert response.status_code == 422


# ---------------------------------------------------------------------
# Swap
# ---------------------------------------------------------------------

def _make_active_reservation(client, headers, setup_id):
    start = datetime.utcnow() + timedelta(hours=1)
    end = start + timedelta(hours=2)
    response = client.post(
        "{0}/reservations".format(API),
        json={"setup_id": setup_id, "reserved_from": _iso(start), "reserved_until": _iso(end)},
        headers=headers,
    )
    assert response.status_code == 201
    return response.json()


def test_swap_request_then_approve(client, auth_headers, developer_user, make_user, make_setup, product):
    dev_headers = auth_headers(developer_user)
    setup_a = make_setup(product_id=product.id)
    setup_b = make_setup(product_id=product.id)
    reservation = _make_active_reservation(client, dev_headers, setup_a.id)

    swap_resp = client.post(
        "{0}/swaps".format(API),
        json={"reservation_id": reservation["id"], "requested_setup_id": setup_b.id},
        headers=dev_headers,
    )
    assert swap_resp.status_code == 201
    swap_id = swap_resp.json()["id"]
    assert swap_resp.json()["status"] == "PENDING"

    approver = make_user(role_name=RoleName.LEAD)
    approve_resp = client.patch(
        "{0}/swaps/{1}/approve".format(API, swap_id), json={}, headers=auth_headers(approver)
    )
    assert approve_resp.status_code == 200
    assert approve_resp.json()["status"] == "COMPLETED"


def test_swap_to_unavailable_setup_rejected(client, auth_headers, developer_user, make_setup, product):
    dev_headers = auth_headers(developer_user)
    setup_a = make_setup(product_id=product.id)
    setup_b = make_setup(product_id=product.id, status=SetupStatus.MAINTENANCE)
    reservation = _make_active_reservation(client, dev_headers, setup_a.id)

    response = client.post(
        "{0}/swaps".format(API),
        json={"reservation_id": reservation["id"], "requested_setup_id": setup_b.id},
        headers=dev_headers,
    )
    assert response.status_code == 409


def test_swap_other_users_reservation_rejected(client, auth_headers, developer_user, second_developer_user, make_setup, product):
    dev_headers = auth_headers(developer_user)
    setup_a = make_setup(product_id=product.id)
    setup_b = make_setup(product_id=product.id)
    reservation = _make_active_reservation(client, dev_headers, setup_a.id)

    response = client.post(
        "{0}/swaps".format(API),
        json={"reservation_id": reservation["id"], "requested_setup_id": setup_b.id},
        headers=auth_headers(second_developer_user),
    )
    assert response.status_code == 403


# ---------------------------------------------------------------------
# Unreserve
# ---------------------------------------------------------------------

def test_unreserve_own_reservation(client, auth_headers, developer_user, setup):
    dev_headers = auth_headers(developer_user)
    reservation = _make_active_reservation(client, dev_headers, setup.id)

    response = client.patch("{0}/reservations/{1}/cancel".format(API, reservation["id"]), headers=dev_headers)
    assert response.status_code == 200
    assert response.json()["status"] == "CANCELLED"


def test_unreserve_another_users_reservation_rejected(client, auth_headers, developer_user, second_developer_user, setup):
    reservation = _make_active_reservation(client, auth_headers(developer_user), setup.id)

    response = client.patch(
        "{0}/reservations/{1}/cancel".format(API, reservation["id"]), headers=auth_headers(second_developer_user)
    )
    assert response.status_code == 403


def test_unreserve_blocked_while_swap_pending(client, auth_headers, developer_user, make_user, make_setup, product):
    dev_headers = auth_headers(developer_user)
    setup_a = make_setup(product_id=product.id)
    setup_b = make_setup(product_id=product.id)
    reservation = _make_active_reservation(client, dev_headers, setup_a.id)

    swap_resp = client.post(
        "{0}/swaps".format(API),
        json={"reservation_id": reservation["id"], "requested_setup_id": setup_b.id},
        headers=dev_headers,
    )
    assert swap_resp.status_code == 201

    cancel_resp = client.patch("{0}/reservations/{1}/cancel".format(API, reservation["id"]), headers=dev_headers)
    assert cancel_resp.status_code == 409


# ---------------------------------------------------------------------
# Announcements
# ---------------------------------------------------------------------

def test_create_announcement_requires_manage_permission(client, auth_headers, developer_user):
    response = client.post(
        "{0}/announcements".format(API),
        json={
            "title": "Blocked",
            "message": "Should not be allowed",
            "start_date": _iso(datetime.utcnow()),
        },
        headers=auth_headers(developer_user),
    )
    assert response.status_code == 403


def test_dev_lead_can_create_and_list_announcement(client, auth_headers, make_user):
    dev_lead = make_user(role_name=RoleName.DEVELOPER_LEAD)
    headers = auth_headers(dev_lead)

    create = client.post(
        "{0}/announcements".format(API),
        json={"title": "System Maintenance", "message": "Downtime tonight", "start_date": _iso(datetime.utcnow())},
        headers=headers,
    )
    assert create.status_code == 201

    listing = client.get("{0}/announcements".format(API), headers=headers)
    assert listing.status_code == 200
    assert listing.json()["total_items"] >= 1


# ---------------------------------------------------------------------
# Logging (audit trail)
# ---------------------------------------------------------------------

def test_login_action_is_audit_logged(client, make_user, auth_headers):
    make_user(username="auditlogin", password="Password123")
    client.post("{0}/auth/login".format(API), json={"username": "auditlogin", "password": "Password123"})

    dev_lead = make_user(role_name=RoleName.DEVELOPER_LEAD)
    listing = client.get(
        "{0}/audit-logs?action=LOGIN".format(API), headers=auth_headers(dev_lead)
    )
    assert listing.status_code == 200
    assert listing.json()["total_items"] >= 1


def test_reservation_create_is_audit_logged(client, auth_headers, developer_user, make_user, setup):
    _make_active_reservation(client, auth_headers(developer_user), setup.id)

    dev_lead = make_user(role_name=RoleName.DEVELOPER_LEAD)
    listing = client.get(
        "{0}/audit-logs?entity_type=Reservation&action=CREATE".format(API), headers=auth_headers(dev_lead)
    )
    assert listing.status_code == 200
    assert listing.json()["total_items"] >= 1
