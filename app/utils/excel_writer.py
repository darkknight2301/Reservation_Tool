"""
Excel writer utilities (openpyxl).

Produces self-describing workbooks: a data sheet with a bold, frozen header
row and auto-sized columns, plus a second "Export Info" sheet documenting
who ran the export, when, and with what filters -- so a downloaded file
remains meaningful even out of the context of the application.
"""
from datetime import datetime
from typing import Any, Dict, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")


def _autosize_columns(worksheet: Worksheet, headers: Sequence[str]) -> None:
    """Set a reasonable column width based on header and content length."""
    for index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(index)
        max_length = len(str(header))
        for cell in worksheet[column_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 4, 60)


def write_excel_workbook(
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    sheet_name: str,
    export_context: Dict[str, Any],
    file_path: str,
) -> None:
    """
    Write a workbook with one data sheet and one "Export Info" metadata sheet.

    Args:
        headers: column header labels, in order.
        rows: one sequence of cell values per data row, matching header order.
        sheet_name: name for the primary data worksheet.
        export_context: metadata (exported_by, exported_at, filters, row_count)
            rendered onto the second sheet.
        file_path: absolute/relative path to write the ``.xlsx`` file to.
    """
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = sheet_name[:31]  # Excel sheet name length limit

    data_sheet.append(list(headers))
    for column_index in range(1, len(headers) + 1):
        cell = data_sheet.cell(row=1, column=column_index)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    data_sheet.freeze_panes = "A2"

    for row in rows:
        data_sheet.append(list(row))

    _autosize_columns(data_sheet, headers)

    info_sheet = workbook.create_sheet(title="Export Info")
    info_sheet.append(["Field", "Value"])
    info_sheet["A1"].font = _HEADER_FONT
    info_sheet["B1"].font = _HEADER_FONT
    info_sheet["A1"].fill = _HEADER_FILL
    info_sheet["B1"].fill = _HEADER_FILL
    for key, value in export_context.items():
        info_sheet.append([str(key), str(value)])
    info_sheet.column_dimensions["A"].width = 25
    info_sheet.column_dimensions["B"].width = 60

    workbook.save(file_path)


def build_export_context(
    exported_by: str,
    exported_at: datetime,
    filters: Dict[str, Any],
    row_count: int,
) -> Dict[str, Any]:
    """Build the metadata dict rendered onto an export's 'Export Info' sheet."""
    return {
        "Exported By": exported_by,
        "Exported At (UTC)": exported_at.isoformat(),
        "Filters Applied": filters if filters else "None",
        "Row Count": row_count,
    }
