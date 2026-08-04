"""
Excel reader utilities (openpyxl) for the Setup bulk-import workflow.

Parsing is deliberately strict: the header row must match the expected
column contract exactly (order-independent, but every expected column must
be present), and rows are returned as plain dicts so the caller (the
ImportService) can validate them through the very same Pydantic schemas
used by the JSON API -- guaranteeing import and API validation never drift
apart.
"""
from typing import Any, Dict, List, Sequence

from openpyxl import load_workbook

from app.core.exceptions import ImportValidationError

# Column headers expected in a Setup import workbook, in the canonical order
# used when generating the import template. Import parsing itself is
# order-independent (matched by header name), but every one of these must
# be present.
SETUP_IMPORT_COLUMNS: List[str] = [
    "product_name",
    "group_name",
    "ip_address",
    "hostname",
    "ssd",
    "hdd",
    "hardware_info",
    "capacity",
    "form_factor",
    "owner_username",
    "adapter",
    "aardvark",
    "quarch",
    "apc",
    "remote_server",
    "location",
    "remarks",
]


def read_setup_import_rows(file_path: str) -> List[Dict[str, Any]]:
    """
    Parse an uploaded Setup import workbook into a list of row dicts.

    Raises:
        ImportValidationError: if the header row is missing required
            columns.
    """
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]

    rows_iterator = worksheet.iter_rows(values_only=True)
    try:
        header_row: Sequence[Any] = next(rows_iterator)
    except StopIteration:
        raise ImportValidationError("The uploaded workbook is empty.")

    header_labels = [str(cell).strip() if cell is not None else "" for cell in header_row]
    missing_columns = [col for col in SETUP_IMPORT_COLUMNS if col not in header_labels]
    if missing_columns:
        raise ImportValidationError(
            "The uploaded workbook is missing required column(s): {0}".format(", ".join(missing_columns))
        )

    column_index_map = {label: index for index, label in enumerate(header_labels)}

    parsed_rows: List[Dict[str, Any]] = []
    for row_values in rows_iterator:
        if row_values is None or all(value is None for value in row_values):
            continue  # skip fully blank rows
        row_dict: Dict[str, Any] = {}
        for column_name in SETUP_IMPORT_COLUMNS:
            index = column_index_map[column_name]
            value = row_values[index] if index < len(row_values) else None
            row_dict[column_name] = value.strip() if isinstance(value, str) else value
        parsed_rows.append(row_dict)

    return parsed_rows
