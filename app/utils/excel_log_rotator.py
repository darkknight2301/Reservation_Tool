"""
Rotating Excel transaction log writer.

Every Excel import/export transaction is additionally appended to a rolling
set of workbooks on disk, organized as::

    {EXCEL_LOG_DIR}/{Month_Year}/{timestamp}_{00001}.xlsx
    {EXCEL_LOG_DIR}/{Month_Year}/{timestamp}_{00002}.xlsx
    ...

A new file is started whenever the current one would exceed
``EXCEL_LOG_ROTATE_MAX_BYTES`` (default 40MB). This is distinct from the
``export_logs``/``excel_transaction_logs`` database tables (which remain the
queryable source of truth); these workbooks exist so the Developer Logs
screen can present and download the raw Excel transaction trail, per the
required ``Logs/Month_Year/Timestamp_00001.xlsx`` layout.
"""
import glob
import os
import re
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook

from app.core.config import settings

_HEADERS = ["timestamp", "operation", "entity_type", "row_number", "status", "message", "user"]
_FILENAME_RE = re.compile(r"^(?P<timestamp>\d{8}_\d{6})_(?P<seq>\d{5})\.xlsx$")


def _month_year_folder(as_of: datetime) -> str:
    """e.g. 'August_2026' -- matches the required Month_Year folder naming."""
    return as_of.strftime("%B_%Y")


def _ensure_month_folder(as_of: datetime) -> str:
    folder = os.path.join(settings.EXCEL_LOG_DIR, _month_year_folder(as_of))
    os.makedirs(folder, exist_ok=True)
    return folder


def _latest_log_file(folder: str) -> str:
    """Return the highest-sequence .xlsx file in the folder, or '' if none exists yet."""
    candidates = []
    for path in glob.glob(os.path.join(folder, "*.xlsx")):
        match = _FILENAME_RE.match(os.path.basename(path))
        if match:
            candidates.append((int(match.group("seq")), path))
    if not candidates:
        return ""
    candidates.sort(key=lambda item: item[0])
    return candidates[-1][1]


def _next_sequence_number(folder: str) -> int:
    existing = [
        int(_FILENAME_RE.match(os.path.basename(p)).group("seq"))
        for p in glob.glob(os.path.join(folder, "*.xlsx"))
        if _FILENAME_RE.match(os.path.basename(p))
    ]
    return (max(existing) + 1) if existing else 1


def _create_new_log_file(folder: str, as_of: datetime) -> str:
    seq = _next_sequence_number(folder)
    filename = "{0}_{1:05d}.xlsx".format(as_of.strftime("%Y%m%d_%H%M%S"), seq)
    path = os.path.join(folder, filename)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"
    sheet.append(_HEADERS)
    workbook.save(path)
    return path


def append_excel_transactions(entries: List[Dict[str, Any]]) -> str:
    """
    Append one or more transaction rows to the current rotating log file,
    creating the Month_Year folder and/or a new sequence file as needed.

    Args:
        entries: each dict should have keys matching ``_HEADERS``.

    Returns:
        The path of the workbook the rows were written to.
    """
    if not entries:
        return ""

    as_of = datetime.utcnow()
    folder = _ensure_month_folder(as_of)
    current_path = _latest_log_file(folder)

    if not current_path or os.path.getsize(current_path) >= settings.EXCEL_LOG_ROTATE_MAX_BYTES:
        current_path = _create_new_log_file(folder, as_of)

    workbook = load_workbook(current_path)
    sheet = workbook["Transactions"] if "Transactions" in workbook.sheetnames else workbook.active

    for entry in entries:
        sheet.append([entry.get(header, "") for header in _HEADERS])

    workbook.save(current_path)

    # Re-check size after writing; if this write pushed us over the limit,
    # the *next* call will rotate to a fresh file (rows already written here
    # are not retroactively split, which keeps this operation simple and atomic).
    return current_path
